import unittest
import tempfile
import json
import csv
from contextlib import closing
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from news_pipeline.config import load_config
from news_pipeline.services.exporter import ExporterService

CONFIG_PATH = Path(__file__).resolve().parents[1] / "config.json"


def _make_test_db(db_path, rows):
    """임시 SQLite DB를 만들고 clean_news에 테스트 데이터를 넣는다."""
    from news_pipeline.database import Database

    db = Database(str(db_path))
    with closing(db.get_connection()) as conn:
        conn.execute("PRAGMA foreign_keys = OFF")
        for row in rows:
            conn.execute(
                """
                INSERT INTO clean_news (
                    raw_id, source, category, title, canonical_url, published_at,
                    content, summary, key_points, summary_status, summarized_at,
                    ai_provider, ai_model, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["raw_id"], row["source"], row["category"], row["title"],
                    row["canonical_url"], row.get("published_at"), row.get("content"),
                    row.get("summary"), row.get("key_points", "[]"),
                    row.get("summary_status", "pending"), row.get("summarized_at"),
                    row.get("ai_provider"), row.get("ai_model"),
                    row.get("created_at", "2026-08-26T00:00:00"),
                    row.get("updated_at", "2026-08-26T00:00:00"),
                )
            )
        conn.commit()
    return db_path

class ExporterServiceTestCase(unittest.TestCase):
    def setUp(self):
        self.service = ExporterService()
        self.tmp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.tmp_dir.name) / "test.db"
        _make_test_db(
            self.db_path,
            rows=[
                {
                    "raw_id": 1,
                    "source": "inews24",
                    "category": "it",
                    "title": "AI 반도체 시장 급성장",
                    "canonical_url": "https://example.com/1",
                    "published_at": "2026-08-20",
                    "summary": "AI 반도체 수요가 늘고 있다.",
                    "key_points": '["수요 증가", "가격 상승"]',
                    "summary_status": "summarized",
                    "summarized_at": "2026-08-20T10:00:00",
                },
                {
                    "raw_id": 2,
                    "source": "inews24",
                    "category": "economy",
                    "title": "금리 동결 발표",
                    "canonical_url": "https://example.com/2",
                    "published_at": "2026-08-21",
                    "summary_status": "pending",
                },
            ],
        )

        from news_pipeline.database import Database
        db = Database(str(self.db_path))
        rows = db.list_news(limit=None)
        for row in rows:
            key_points = row.get("key_points")
            if isinstance(key_points, str):
                try:
                    row["key_points"] = json.loads(key_points)
                except (json.JSONDecodeError, TypeError):
                    row["key_points"] = []
        self.news = [
            {col: row.get(col) for col in self.service.EXPORT_COLUMNS} for row in rows
        ]

    def tearDown(self):
        self.tmp_dir.cleanup()

    def test_filter_status_summarized_only(self):
        filtered = self.service._filter_clean_news(
            self.news, status="summarized", category=None, date_from=None, date_to=None
        )
        self.assertTrue(all(n["summary_status"] == "summarized" for n in filtered))

    def test_filter_status_unsummarized_excludes_summarized(self):
        filtered = self.service._filter_clean_news(
            self.news, status="unsummarized", category=None, date_from=None, date_to=None
        )
        self.assertTrue(all(n["summary_status"] != "summarized" for n in filtered))

    def test_filter_by_category(self):
        filtered = self.service._filter_clean_news(
            self.news, status="all", category="it", date_from=None, date_to=None
        )
        self.assertTrue(all(n["category"] == "it" for n in filtered))

    def test_filter_category_all_returns_every_category(self):
        """--category all은 특정 카테고리로 좁히지 않고 전체를 반환해야 한다."""
        filtered = self.service._filter_clean_news(
            self.news, status="all", category="all", date_from=None, date_to=None
        )
        self.assertEqual(len(filtered), len(self.news))

    def test_filter_by_date_range(self):
        filtered = self.service._filter_clean_news(
            self.news, status="all", category=None,
            date_from="2026-08-21", date_to="2026-08-21",
        )
        self.assertTrue(all(n["published_at"] == "2026-08-21" for n in filtered))

    def test_filter_combined_returns_empty_when_no_match(self):
        filtered = self.service._filter_clean_news(
            self.news, status="summarized", category="economy", date_from=None, date_to=None
        )
        self.assertEqual(filtered, [])

    def test_filter_date_range_excludes_none_published_at_without_error(self):
        """published_at이 None인 뉴스는 날짜 필터 적용 시 에러 없이 제외되어야 한다."""
        news_with_missing_date = self.news + [
            {**self.news[0], "id": 999, "title": "발행일 없음", "published_at": None}
        ]
        filtered = self.service._filter_clean_news(
            news_with_missing_date, status="all", category=None,
            date_from="2026-08-01", date_to="2026-08-31",
        )
        self.assertNotIn(None, [n["published_at"] for n in filtered])
        self.assertEqual(len(filtered), len(self.news))

    def test_filter_date_range_ignores_time_component(self):
        """published_at에 시간이 포함돼도 date_to 당일 기사가 제외되지 않아야 한다."""
        news_with_time = self.news + [
            {**self.news[0], "id": 998, "title": "시간 포함 발행일", "published_at": "2026-08-21T23:59:00"}
        ]
        filtered = self.service._filter_clean_news(
            news_with_time, status="all", category=None,
            date_from="2026-08-21", date_to="2026-08-21",
        )
        titles = {n["title"] for n in filtered}
        self.assertIn("시간 포함 발행일", titles)

    def test_save_csv_uses_utf8_sig_encoding(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test.csv"
            self.service._save_csv(self.news, output_path)
            with open(output_path, "rb") as f:
                raw = f.read()
            self.assertTrue(raw.startswith(b"\xef\xbb\xbf"))

    def test_save_jsonl_produces_valid_json_lines(self):
        import json
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test.jsonl"
            self.service._save_jsonl(self.news, output_path)
            with open(output_path, encoding="utf-8") as f:
                lines = f.readlines()
            self.assertEqual(len(lines), len(self.news))
            for line in lines:
                json.loads(line)

    def test_save_xlsx_has_correct_headers(self):
        from openpyxl import load_workbook
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test.xlsx"
            self.service._save_xlsx(self.news, output_path)
            wb = load_workbook(output_path)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(tuple(headers), self.service.EXPORT_COLUMNS)

    def test_save_xlsx_wraps_long_title_and_summary_cells(self):
        """제목·요약처럼 길어질 수 있는 컬럼은 줄바꿈이 적용되고 폭이 과도하게 늘어나지 않아야 한다."""
        from openpyxl import load_workbook

        long_news = [
            {
                **self.news[0],
                "title": "가" * 200,
                "summary": "나" * 300,
            }
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test.xlsx"
            self.service._save_xlsx(long_news, output_path)
            wb = load_workbook(output_path)
            sheet = wb.active

            title_col_index = self.service.EXPORT_COLUMNS.index("title") + 1
            summary_col_index = self.service.EXPORT_COLUMNS.index("summary") + 1
            title_letter = sheet.cell(row=1, column=title_col_index).column_letter
            summary_letter = sheet.cell(row=1, column=summary_col_index).column_letter

            title_cell = sheet.cell(row=2, column=title_col_index)
            summary_cell = sheet.cell(row=2, column=summary_col_index)

            self.assertTrue(title_cell.alignment.wrap_text)
            self.assertTrue(summary_cell.alignment.wrap_text)
            self.assertLessEqual(sheet.column_dimensions[title_letter].width, 62)
            self.assertLessEqual(sheet.column_dimensions[summary_letter].width, 62)

    def test_save_csv_handles_commas_and_newlines(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "comma_test.db"
            _make_test_db(
                db_path,
                rows=[
                    {
                        "raw_id": 99,
                        "source": "inews24",
                        "category": "it",
                        "title": "쉼표, 줄바꿈\n테스트 제목",
                        "canonical_url": "https://example.com/99",
                        "published_at": "2026-08-22",
                        "summary": "요약에, 쉼표와\n줄바꿈이 포함된 경우",
                        "key_points": '["포인트 A, B", "줄바꿈\\n포함"]',
                        "summary_status": "summarized",
                    },
                ],
            )
            from news_pipeline.database import Database
            db = Database(str(db_path))
            rows = db.list_news(limit=10)
            for row in rows:
                key_points = row.get("key_points")
                if isinstance(key_points, str):
                    row["key_points"] = json.loads(key_points)
            news = [
                {col: row.get(col) for col in self.service.EXPORT_COLUMNS} for row in rows
            ]

            output_path = Path(tmpdir) / "comma_test.csv"
            self.service._save_csv(news, output_path)

            with open(output_path, encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                result_rows = list(reader)

            self.assertEqual(len(result_rows), 1)
            self.assertEqual(result_rows[0]["title"], "쉼표, 줄바꿈\n테스트 제목")
            self.assertEqual(
                result_rows[0]["summary"], "요약에, 쉼표와\n줄바꿈이 포함된 경우"
            )


class ExporterServiceIntegrationTestCase(unittest.TestCase):
    """DB 삽입부터 ExporterService.export() 전체 흐름(조회→필터→저장)을 직접 검증한다."""

    def setUp(self):
        self.tmp_dir = tempfile.TemporaryDirectory()
        self.project_root = Path(self.tmp_dir.name)
        self.db_path = self.project_root / "data" / "news.db"
        _make_test_db(
            self.db_path,
            rows=[
                {
                    "raw_id": 1,
                    "source": "inews24",
                    "category": "it",
                    "title": "AI 반도체 시장 급성장",
                    "canonical_url": "https://example.com/1",
                    "published_at": "2026-08-20",
                    "summary": "AI 반도체 수요가 늘고 있다.",
                    "key_points": '["수요 증가", "가격 상승"]',
                    "summary_status": "summarized",
                    "summarized_at": "2026-08-20T10:00:00",
                },
                {
                    "raw_id": 2,
                    "source": "inews24",
                    "category": "economy",
                    "title": "금리 동결 발표",
                    "canonical_url": "https://example.com/2",
                    "published_at": "2026-08-21",
                    "summary_status": "pending",
                },
            ],
        )

        base_config, _ = load_config(CONFIG_PATH)
        self.export_dir = self.project_root / "exports"
        self.config = base_config.model_copy(
            update={
                "database": base_config.database.model_copy(update={"path": str(self.db_path)}),
                "export": base_config.export.model_copy(
                    update={"output_directory": str(self.export_dir)}
                ),
            }
        )
        self.service = ExporterService()

    def tearDown(self):
        self.tmp_dir.cleanup()

    def _export(self, **overrides):
        params = dict(
            output_format="csv",
            status="all",
            category=None,
            date_from=None,
            date_to=None,
            output=None,
            config=self.config,
            project_root=self.project_root,
        )
        params.update(overrides)
        return self.service.export(**params)

    def test_export_full_flow_writes_expected_csv_rows(self):
        """mock 없이 실제 DB 조회 → 필터 → CSV 저장까지 전체 흐름을 검증한다."""
        result_path = self._export(output_format="csv")

        self.assertTrue(result_path.exists())
        with open(result_path, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f))

        self.assertEqual(len(rows), 2)
        titles = {row["title"] for row in rows}
        self.assertEqual(titles, {"AI 반도체 시장 급성장", "금리 동결 발표"})

    def test_export_full_flow_applies_status_filter(self):
        """export()에 전달한 status 필터가 실제 DB 조회 결과에 반영되는지 확인한다."""
        result_path = self._export(output_format="jsonl", status="summarized")

        with open(result_path, encoding="utf-8") as f:
            rows = [json.loads(line) for line in f]

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["summary_status"], "summarized")

    def test_export_same_second_consecutive_runs_produce_different_files(self):
        """같은 시각에 연속 실행해도 파일이 서로 다르며 기존 파일을 덮어쓰지 않는다."""
        fixed_now = datetime(2026, 8, 26, 12, 0, 0, 123456)
        with patch("news_pipeline.services.exporter.datetime") as mock_datetime:
            mock_datetime.now.return_value = fixed_now
            first_path = self._export(output_format="csv")
            second_path = self._export(output_format="csv")

        self.assertNotEqual(first_path, second_path)
        self.assertTrue(first_path.exists())
        self.assertTrue(second_path.exists())
        with open(first_path, encoding="utf-8-sig", newline="") as f:
            self.assertEqual(len(list(csv.DictReader(f))), 2)
        with open(second_path, encoding="utf-8-sig", newline="") as f:
            self.assertEqual(len(list(csv.DictReader(f))), 2)
